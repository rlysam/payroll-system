#library to import the excel file
import openpyxl
#libraries to create the pdf file and add text to it
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
from functions.pdf_generator.forExcel import *
#libraries to merge pdf files
import os
from PyPDF2 import PdfFileReader, PdfFileMerger

#convert the font so it is compatible
pdfmetrics.registerFont(TTFont('Ayar','functions/pdf_generator/ayar.ttf'))
# pdfmetrics.registerFont(TTFont('Ayar','ayar.ttf'))

#Page information
page_width = 2156
page_height = 3050
spread = 100
start = 200
start_2 = 700

#Payslip variables
company_name = 'SGV Co./EY Philippines'
month_year = 'January 2022'

def create_payslip(mapData):
    dfAndSaveToExcel(mapData)
    #import the sheet from the excel file
    wb = openpyxl.load_workbook('functions/pdf_generator/data/data_payslip.xlsx')
    # wb = openpyxl.load_workbook('data/data_payslip.xlsx')
    sheet = wb.get_sheet_by_name('employees')

    for i in range (2, 42):
        #eading values from excel file
        pay_date = sheet.cell(row = i, column = 1).value
        employee_name = sheet.cell(row = i, column = 2).value
        basic_salary = sheet.cell(row = i, column = 12).value
        overtime = sheet.cell(row = i, column = 13).value
        deminimis = sheet.cell(row = i, column = 11).value
        gross_pay = sheet.cell(row = i, column = 14).value

        tax = sheet.cell(row = i, column = 15).value
        sss = sheet.cell(row = i, column = 16).value
        philhealth_payment = sheet.cell(row = i, column = 17).value
        hdmf = sheet.cell(row = i, column = 18).value
        deductions = sheet.cell(row = i, column = 19).value

        net_pay = sheet.cell(row = i, column = 20).value

        #Creating a pdf file and setting a naming convention
        c = canvas.Canvas(str(employee_name)+'_PAYSLIP'+'.pdf' )
        # c = canvas.Canvas('_PAYSLIP'+'.pdf' )
        # c = canvas.Canvas('email/'+str(employee_name)+'.pdf' )
        #Page settings (size/font)
        c.setPageSize((page_width, page_height))
        c.setFont('Ayar',80)

        #Company name text
        text_width = stringWidth(company_name, 'Ayar',80)
        c.drawString((page_width-text_width)/2, 2900, company_name)
        #Invoice month/year information
        text = 'Salary calculation for period ' + month_year
        text_width = stringWidth(text, 'Ayar',55)
        c.setFont('Ayar',55)
        c.drawString((page_width-text_width)/2, 2700, text)

        y = 2500
        #Drawing payslip related information
        c.setFont('Ayar',45)
        c.drawString(start, y, 'Pay Date - dd/mm/yy :')
        c.drawString(start_2, y, str(pay_date))
        y -= spread

        c.drawString(start, y, 'Employee name:')
        c.drawString(start_2, y, str(employee_name) + ' ')
        y -= spread

        c.drawString(start, y, 'Basic Salary:')
        c.drawString(start_2, y, str(basic_salary))
        y -= spread

        c.drawString(start, y, 'De Minimis:')
        c.drawString(start_2, y, str(deminimis))
        y -= spread

        c.drawString(start, y, 'Overtime:')
        c.drawString(start_2, y, str(overtime))
        y -= spread

        c.drawString(start, y, 'Gross Pay:')
        c.drawString(start_2, y, str(gross_pay))
        y -= spread

        c.drawString(start, y, ' ')
        y -= spread

        c.drawString(start, y, 'Tax:')
        c.drawString(start_2, y, str(tax))
        y -= spread

        c.drawString(start, y, 'SSS:')
        c.drawString(start_2, y, str(sss))
        y -= spread

        c.drawString(start, y, 'PhilHealth Payment:')
        c.drawString(start_2, y, str(philhealth_payment))
        y -= spread

        c.drawString(start, y, 'HDMF (Pag-IBIG):')
        c.drawString(start_2, y, str(hdmf))
        y -= spread

        c.drawString(start, y, 'Deductions:')
        c.drawString(start_2, y, str(deductions))
        y -= spread

        c.drawString(start, y, ' ')
        y -= spread

        c.drawString(start, y, 'Net Pay:')
        c.drawString(start_2, y, str(net_pay))
        y -= spread

        c.drawString(start, y, ' ')
        y -= spread

        c.drawString(start, y, 'Signature: ')
        c.drawString(start_2, y,'____________')

        #Saving the pdf file
        c.save()

# # basura ata to...
# def merge_pdfs():
#     files_dir = 'tutorial' #Select the directory where the pdf files are located
#     pdf_files = [f for f in os.listdir(files_dir) if f.endswith('.pdf')] #Get all files in the directory that end with '.pdf'
#     merger = PdfFileMerger() #Create an empty file
#     for filename in pdf_files:
#         merger.append(PdfFileReader(os.path.join(files_dir,filename),'rb')) #Add every pdf to the empty file
#     merger.write(os.path.join(files_dir,'merged_pdfs.pdf')) #Save the file

# a={
# 	"deductions": "₱2258.54",
# 	"gross_pay": "₱26502.00",
# 	"hdmf": "₱100.00",
# 	"net_pay": "₱24243.46",
# 	"other_payment_due": "0",
# 	"pension_pay": "₱1192.59",
# 	"philhealth_payment": "₱88.34",
# 	"sss": "₱1192.59",
# 	"tax": "₱857.61",
# 	"taxable_pay": "₱857.61",
# 	"pay_date": "AMOGAS Data 1",
# 	"employee_name": "BELLO",
# 	"address": "Sample Data 1",
# 	"reference": "Sample Data 1",
# 	"employer_name": "Sample Data 1",
# 	"email": "sam17.bello@ymail.com",
# 	"job_status": "Sample data 2",
# 	"postcode": "Sample data 2",
# 	"grade": "Sample data 2",
# 	"department": "Sample data 2",
# 	"deminimis": "Sample data 2",
# 	"basic_salary": "PHP 20000",
# 	"overtime": "OVERTIME_FROM_PDF_GEN",
# 	"loan": "Sample data 2",
# 	"philhealth_number": "Sample data 2"
# }

# create_payslip(a)
# merge_pdfs()
