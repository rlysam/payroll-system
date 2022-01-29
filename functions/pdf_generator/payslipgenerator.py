#library to import the excel file
import openpyxl
#libraries to create the pdf file and add text to it
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
#libraries to merge pdf files
import os
from PyPDF2 import PdfFileReader, PdfFileMerger


#import the sheet from the excel file
# wb = openpyxl.load_workbook('data/data_payslip.xlsx')
# wb = openpyxl.load_workbook('data/data_payslip.xlsx')
wb = openpyxl.load_workbook('functions/pdf_generator/data/data_payslip.xlsx')
sheet = wb.get_sheet_by_name('employees')

#convert the font so it is compatible
pdfmetrics.registerFont(TTFont('Ayar','functions/pdf_generator/ayar.ttf'))

#Page information
page_width = 2156
page_height = 3050
spread = 100
start = 200
start_2 = 700

#Payslip variables
company_name = 'SGV Co./EY Philippines'
month_year = 'January 2022'

def create_payslip():
    for i in range (2, 42):
        #eading values from excel file
        employee_name = sheet.cell(row = i, column = 1).value
        address = sheet.cell(row = i, column = 2).value
        reference = sheet.cell(row = i, column = 3).value
        employer_name = sheet.cell(row = i, column = 4).value
        email = sheet.cell(row = i, column = 5).value
        job_status = sheet.cell(row = i, column = 6).value
        postcode = sheet.cell(row = i, column = 7).value
        grade = sheet.cell(row = i, column = 8).value
        department = sheet.cell(row = i, column = 9).value
        deminimis = sheet.cell(row = i, column = 10).value
        basic_salary = sheet.cell(row = i, column = 11).value
        overtime = sheet.cell(row = i, column = 12).value
        gross_pay = sheet.cell(row = i, column = 13).value
        net_pay = sheet.cell(row = i, column = 14).value
        tax = sheet.cell(row = i, column = 15).value
        sss = sheet.cell(row = i, column = 16).value
        loan = sheet.cell(row = i, column = 17).value
        philhealth_payment = sheet.cell(row = i, column = 18).value
        hdmf = sheet.cell(row = i, column = 19).value
        deductions = sheet.cell(row = i, column = 20).value
        pay_date = sheet.cell(row = i, column = 21).value
        philhealth_number = sheet.cell(row = i, column = 22).value
        taxable_pay = sheet.cell(row = i, column = 23).value
        pension_pay = sheet.cell(row = i, column = 24).value
        other_payment_due = sheet.cell(row = i, column = 25).value

        #Creating a pdf file and setting a naming convention
        c = canvas.Canvas('functions/pdf_generator/data/'+str(employee_name)+'_'+ month_year + '.pdf' )
        #Page settings (size/font)
        c.setPageSize((page_width, page_height))
        c.setFont('Ayar',80)
        #Company name text
        text_width = stringWidth(company_name, 'Ayar',80)
        c.drawString((page_width-text_width)/2, 2900, company_name)
        #Invoice month/year information
        text = 'Salary calculation for period ' + month_year
        text_width = stringWidth(text, 'Ayar',55)
        c.setFont('Ayar',55)
        c.drawString((page_width-text_width)/2, 2700, text)

        y = 2500
        #Drawing payslip related information
        c.setFont('Ayar',45)
        c.drawString(start, y, 'Employee\'s name:')
        c.drawString(start_2, y, str(employee_name) + ' ')
        y -= spread

        c.drawString(start, y, 'Address:')
        c.drawString(start_2, y, str(address))
        y -= spread

        c.drawString(start, y, 'Reference:')
        c.drawString(start_2, y, str(reference))
        y -= spread

        c.drawString(start, y, 'Employer Name:')
        c.drawString(start_2, y, str(employer_name))
        y -= spread

        c.drawString(start, y, 'Email:')
        c.drawString(start_2, y, str(email))
        y -= spread
        
        c.drawString(start, y, 'Job Status:')
        c.drawString(start_2, y, str(job_status))
        y -= spread

        c.drawString(start, y, 'Post Code:')
        c.drawString(start_2, y, str(postcode))
        y -= spread

        c.drawString(start, y, 'Grade:')
        c.drawString(start_2, y, str(grade))
        y -= spread

        c.drawString(start, y, 'Department:')
        c.drawString(start_2, y, str(department))
        y -= spread

        c.drawString(start, y, 'De Minimis:')
        c.drawString(start_2, y, str(deminimis))
        y -= spread

        c.drawString(start, y, 'Basic Salary:')
        c.drawString(start_2, y, str(basic_salary))
        y -= spread

        c.drawString(start, y, 'Overtime:')
        c.drawString(start_2, y, str(overtime))
        y -= spread

        c.drawString(start, y, 'Gross Pay:')
        c.drawString(start_2, y, str(gross_pay))
        y -= spread

        c.drawString(start, y, 'Net Pay:')
        c.drawString(start_2, y, str(net_pay))
        y -= spread

        c.drawString(start, y, 'Tax:')
        c.drawString(start_2, y, str(tax))
        y -= spread

        c.drawString(start, y, 'SSS:')
        c.drawString(start_2, y, str(sss))
        y -= spread

        c.drawString(start, y, 'Loan:')
        c.drawString(start_2, y, str(loan))
        y -= spread

        c.drawString(start, y, 'PhilHealth Payment:')
        c.drawString(start_2, y, str(philhealth_payment))
        y -= spread

        c.drawString(start, y, 'HDMF (Pag-IBIG):')
        c.drawString(start_2, y, str(hdmf))
        y -= spread

        c.drawString(start, y, 'Deductions:')
        c.drawString(start_2, y, str(deductions))
        y -= spread

        c.drawString(start, y, 'Pay Date - dd/mm/yy :')
        c.drawString(start_2, y, str(pay_date))
        y -= spread

        c.drawString(start, y, 'PhilHealth Number:')
        c.drawString(start_2, y, str(philhealth_number))
        y -= spread

        c.drawString(start, y, 'Taxable Pay:')
        c.drawString(start_2, y, str(taxable_pay))
        y -= spread

        c.drawString(start, y, 'SSS Pay:')
        c.drawString(start_2, y, str(pension_pay))
        y -= spread

        c.drawString(start, y, 'Other Payment Due:')
        c.drawString(start_2, y, str(other_payment_due))
        y -= spread

        c.drawString(start, y, 'Net Pay:')
        c.drawString(start_2, y, str(net_pay))
        y -= spread * 3

        c.drawString(start, y, 'Signature: ')
        c.drawString(start_2, y,'____________')
      
        #Saving the pdf file
        c.save()

# # basura ata to...
# def merge_pdfs():
#     files_dir = 'tutorial' #Select the directory where the pdf files are located
#     pdf_files = [f for f in os.listdir(files_dir) if f.endswith('.pdf')] #Get all files in the directory that end with '.pdf'
#     merger = PdfFileMerger() #Create an empty file
#     for filename in pdf_files:
#         merger.append(PdfFileReader(os.path.join(files_dir,filename),'rb')) #Add every pdf to the empty file
#     merger.write(os.path.join(files_dir,'merged_pdfs.pdf')) #Save the file

create_payslip()
# merge_pdfs()
